#!/usr/bin/env python3
"""Create a conservative Mercado Livre → OpenCart catalog correction plan.

This script deliberately never converts a multi-unit listing into a unit price.
Only a spreadsheet listing that represents the exact OpenCart sale unit is used.
"""
from __future__ import annotations

import html
import json
import math
import re
import subprocess
import unicodedata
from collections import Counter
from pathlib import Path

import openpyxl

XLSX = Path("/Users/mac/Downloads/Anuncios-2026_07_15-14_32.xlsx")
OUT_JSON = Path("scripts/ml-sync-matches.json")
OUT_SQL = Path("scripts/ml-sync-update.sql")
OUT_REPORT = Path("scripts/ml-sync-report.txt")
PT_BR_LANGUAGE_ID = 2


def norm(value: object) -> str:
    text = unicodedata.normalize("NFKD", str(value or "").lower())
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.replace("″", '"').replace("”", '"').replace("“", '"').replace("×", "x")
    return re.sub(r"\s+", " ", text.replace('"', "").strip())


def number(value: object) -> float | None:
    if value is None or value == "":
        return None
    try:
        parsed = float(str(value).strip().replace(",", "."))
    except ValueError:
        return None
    return parsed if math.isfinite(parsed) and parsed >= 0 else None


def integer(value: object) -> int:
    parsed = number(value)
    return max(0, int(parsed)) if parsed is not None else 0


def inch_dimensions(text: str) -> tuple[str, ...] | None:
    """Extract the complete inch dimension sequence, including whole-inch lengths."""
    match = re.search(
        r"(\d+/\d+)\s*x\s*(\d+/\d+|\d+)(?:\s*x\s*(\d+/\d+|\d+))?",
        norm(text),
    )
    if not match:
        return None
    return tuple(part for part in match.groups() if part is not None)


def first_fraction(text: str) -> tuple[str, ...] | None:
    match = re.search(r"\b\d+/\d+\b", norm(text))
    return (match.group(0),) if match else None


def bucha_pair(text: str) -> tuple[str, str] | None:
    match = re.search(r"(?:bucha[^\d]*|cone morse\s*)(\d+)\s*x\s*(\d+)", norm(text))
    return (match.group(1), match.group(2)) if match else None


def cm(text: str) -> str | None:
    match = re.search(r"\b(?:cone\s*morse\s*|cm\s*)([2-5])\b", norm(text))
    return match.group(1) if match else None


def tubular_dimensions(text: str) -> tuple[str, str] | None:
    normalized = norm(text)
    match = re.search(r"\b(\d+)\s*x\s*(\d+)\s*mm\b", normalized)
    if not match:
        match = re.search(r"\bcm\s*[2-5]\s+(\d+)\s*x\s*(\d+)\b", normalized)
    return (match.group(1), match.group(2)) if match else None


def material(text: str) -> str | None:
    normalized = norm(text)
    if any(token in normalized for token in ("10%", "cobalto", "10co")):
        return "10co"
    if any(token in normalized for token in ("50%", "hardms", "hardsteel")):
        return "50"
    return None


def product_family(text: str) -> str:
    normalized = norm(text)
    if ("maleta" in normalized or "jogo de pinc" in normalized) and "mandril" in normalized:
        return "maleta-er40"
    if "porta recartilha" in normalized:
        return "recartilha"
    if "porta bits" in normalized:
        return "porta-bits"
    if "porta bedame" in normalized:
        return "porta-bedame"
    if "bits" in normalized:
        return "bits-redondo" if "redondo" in normalized else "bits-quadrado"
    if "bedame" in normalized:
        return "bedame"
    if "bucha" in normalized:
        return "bucha"
    if "ponta rotativa" in normalized or "contra ponto" in normalized:
        return "ponta"
    return "other"


def point_type(text: str) -> str | None:
    normalized = norm(text)
    for kind in ("tubular", "copiadora", "super", "standard"):
        if kind in normalized:
            return kind
    return None


def recart_type(text: str) -> str | None:
    normalized = norm(text)
    for kind in ("triplo", "duplo", "interno", "simples"):
        if kind in normalized:
            return kind
    return None


def mandril(text: str) -> str | None:
    normalized = norm(text).replace(" ", "")
    for code in ("bt40", "iso50", "iso40", "iso30", "cm5", "cm4", "cm3"):
        if code in normalized:
            return code
    return None


def brand(text: str) -> str | None:
    return "winner-steel" if "winner steel" in norm(text) else None


def sale_unit(title: str, family: str) -> tuple[int, str]:
    """Return explicit unit count and its evidence.  Kit ER40 is one product."""
    normalized = norm(title)
    if family == "maleta-er40":
        return 1, "ER40 kit is the OpenCart product type, not a multi-unit offer"
    if "+" in title or normalized.startswith("kit "):
        return 0, "composite kit/bundle"
    match = re.search(r"\((\d+)\s*unidades?\)", normalized)
    if match:
        return int(match.group(1)), "explicit parenthesized unit count"
    match = re.search(r"^(\d+)\s+(?:bits|porta\s+bits|porta\s+recartilha)\b", normalized)
    if match:
        return int(match.group(1)), "explicit leading unit count"
    return 1, "no multi-unit indicator"


def signature(text: str) -> dict[str, object]:
    family = product_family(text)
    return {
        "family": family,
        "fractions": first_fraction(text) if family in {"porta-bits", "recartilha"} else inch_dimensions(text),
        "material": material(text),
        "brand": brand(text),
        "bucha": bucha_pair(text),
        "cm": cm(text),
        "point_type": point_type(text),
        "tubular": tubular_dimensions(text),
        "recart_type": recart_type(text),
        "mandril": mandril(text),
    }


def mysql_json(query: str) -> list[dict[str, object]]:
    command = [
        "docker", "exec", "-i", "miigtools-ecommerce-mysql-1", "mysql",
        "-uroot", "-popencart", "opencart", "-N", "--batch", "--raw",
        "--default-character-set=utf8mb4", "-e", query,
    ]
    output = subprocess.check_output(command, text=True, stderr=subprocess.DEVNULL)
    return [json.loads(line) for line in output.splitlines() if line.strip()]


def load_products() -> list[dict[str, object]]:
    query = f"""
        SELECT JSON_OBJECT(
            'id', p.product_id, 'sku', p.sku, 'model', p.model, 'name', pd.name,
            'description', pd.description, 'price', p.price, 'quantity', p.quantity,
            'status', p.status, 'length', p.length, 'width', p.width,
            'height', p.height, 'weight', p.weight
        )
        FROM ws_product p
        JOIN ws_product_description pd
          ON pd.product_id = p.product_id AND pd.language_id = {PT_BR_LANGUAGE_ID}
        ORDER BY p.product_id
    """
    products = mysql_json(query)
    for product in products:
        product["signature"] = signature(f"{product['name']} {product['sku']} {product['model']}")
    return products


def load_listings() -> list[dict[str, object]]:
    workbook = openpyxl.load_workbook(XLSX, data_only=True, read_only=True)
    worksheet = workbook["Anúncios"]
    headers = [worksheet.cell(1, column).value for column in range(1, worksheet.max_column + 1)]
    listings = []
    for row_number in range(6, worksheet.max_row + 1):
        source = {headers[index]: worksheet.cell(row_number, index + 1).value for index in range(len(headers))}
        title = str(source.get("TITLE") or "").strip()
        price = number(source.get("PRICE"))
        if not title or price is None:
            continue
        family = product_family(title)
        units, unit_evidence = sale_unit(title, family)
        listing = {
            "row": row_number,
            "sku": str(source.get("SKU") or "").strip(),
            "title": title,
            "price": price,
            "quantity": integer(source.get("QUANTITY")),
            "status": str(source.get("STATUS") or "").strip(),
            "description": str(source.get("DESCRIPTION") or "").strip(),
            "height": number(source.get("SHIPPING_HEIGHT")),
            "width": number(source.get("SHIPPING_WIDTH")),
            "length": number(source.get("SHIPPING_DEPTH")),
            "weight": number(source.get("SHIPPING_WEIGHT")),
            "units": units,
            "unit_evidence": unit_evidence,
            "signature": signature(title),
        }
        listings.append(listing)
    return listings


def compatible(product: dict[str, object], listing: dict[str, object]) -> tuple[bool, str]:
    product_signature = product["signature"]
    listing_signature = listing["signature"]
    family = product_signature["family"]
    if family != listing_signature["family"]:
        return False, "different product family"
    if listing["units"] != 1:
        return False, f"listing sells {listing['units']} units ({listing['unit_evidence']})"
    if family in {"bits-quadrado", "bedame", "porta-bedame"}:
        if not product_signature["fractions"] or product_signature["fractions"] != listing_signature["fractions"]:
            return False, "different or missing exact inch dimensions"
        if product_signature["material"] != listing_signature["material"] or not product_signature["material"]:
            return False, "different or missing material grade"
        if product_signature["brand"] and product_signature["brand"] != listing_signature["brand"]:
            return False, "different or missing required Winner Steel brand"
        return True, "same family, every inch dimension, material grade, and required brand"
    if family == "porta-bits":
        if product_signature["fractions"] == listing_signature["fractions"] and product_signature["fractions"]:
            return True, "same porta-bits family and exact inch size"
        return False, "different or missing porta-bits size"
    if family == "bucha":
        if product_signature["bucha"] == listing_signature["bucha"] and product_signature["bucha"]:
            return True, "same bucha family and exact Morse taper pair"
        return False, "different or missing Morse taper pair"
    if family == "ponta":
        same_base = (
            product_signature["point_type"] == listing_signature["point_type"]
            and product_signature["cm"] == listing_signature["cm"]
            and product_signature["point_type"]
            and product_signature["cm"]
        )
        if not same_base:
            return False, "different or missing point type / Morse cone"
        if product_signature["point_type"] == "tubular":
            if product_signature["tubular"] != listing_signature["tubular"] or not product_signature["tubular"]:
                return False, "different or missing tubular dimensions"
        return True, "same point type, Morse cone, and tubular dimensions when applicable"
    if family == "maleta-er40":
        if product_signature["mandril"] == listing_signature["mandril"] and product_signature["mandril"]:
            return True, "same ER40 kit and exact mandrel interface"
        return False, "different or missing ER40 mandrel interface"
    if family == "recartilha":
        if (
            product_signature["recart_type"] == listing_signature["recart_type"]
            and product_signature["fractions"] == listing_signature["fractions"]
            and product_signature["fractions"]
        ):
            return True, "same recartilha type and exact shank size"
        return False, "recartilha listing lacks the needed exact shank size"
    return False, "unsupported family: no safe deterministic identifier"


def rank(listing: dict[str, object]) -> tuple[int, int, int]:
    return (
        1 if listing["status"].casefold() == "ativo" else 0,
        1 if listing["quantity"] > 0 else 0,
        -listing["row"],
    )


def html_description(text: str) -> str:
    paragraphs = [part.strip() for part in re.split(r"\r?\n\s*\r?\n", text.strip()) if part.strip()]
    return "\n".join(f"<p>{html.escape(part).replace(chr(10), '<br>')}</p>" for part in paragraphs)


def sql_string(value: str) -> str:
    return value.replace("\\", "\\\\").replace("'", "''")


def changed_fields(product: dict[str, object], listing: dict[str, object]) -> dict[str, dict[str, object]]:
    target = {
        "price": round(float(listing["price"]), 4),
        "quantity": listing["quantity"],
        "status": 1 if listing["status"].casefold() == "ativo" else 0,
        "length": listing["length"],
        "width": listing["width"],
        "height": listing["height"],
        "weight": listing["weight"],
    }
    changes = {
        field: {"old": product[field], "new": value}
        for field, value in target.items()
        if value is not None and str(product[field]) != str(value)
    }
    if listing["description"]:
        rendered = html_description(str(listing["description"]))
        if product["description"] != rendered:
            changes["pt_br_description"] = {
                "old_length": len(str(product["description"])),
                "new_length": len(rendered),
            }
    return changes


def build_audit(products: list[dict[str, object]], listings: list[dict[str, object]]) -> dict[str, object]:
    applied, skipped, unmatched, used_rows = [], [], [], set()
    candidates_by_product: dict[int, list[tuple[dict[str, object], str]]] = {}
    for product in products:
        candidates = []
        for listing in listings:
            is_compatible, rationale = compatible(product, listing)
            if is_compatible:
                candidates.append((listing, rationale))
        if candidates:
            candidates_by_product[int(product["id"])] = candidates
    for product in products:
        candidates = candidates_by_product.get(int(product["id"]), [])
        if not candidates:
            continue
        candidates.sort(key=lambda item: rank(item[0]), reverse=True)
        selected, rationale = candidates[0]
        used_rows.add(selected["row"])
        applied.append({
            "product_id": product["id"],
            "oc_name": product["name"],
            "oc_sku": product["sku"],
            "source": selected,
            "confidence": "high",
            "rationale": rationale,
            "alternative_source_rows": [candidate["row"] for candidate, _ in candidates[1:]],
            "changes": changed_fields(product, selected),
            "old": {field: product[field] for field in ("price", "quantity", "status", "length", "width", "height", "weight")},
        })
        for alternate, alternate_rationale in candidates[1:]:
            skipped.append({
                "row": alternate["row"], "sku": alternate["sku"], "title": alternate["title"],
                "reason": f"duplicate exact match for product {product['id']}; selected row {selected['row']} because active/in-stock precedence",
                "rationale": alternate_rationale,
            })
    for listing in listings:
        if listing["row"] in used_rows or any(item["row"] == listing["row"] for item in skipped):
            continue
        identity_listing = {**listing, "units": 1}
        exact_identity_products = [
            product for product in products if compatible(product, identity_listing)[0]
        ]
        possible_products = []
        for product in products:
            product_signature = product["signature"]
            listing_signature = listing["signature"]
            if product_signature["family"] == listing_signature["family"] and product_signature["family"] != "other":
                possible_products.append(product)
        if listing["units"] != 1:
            item = {
                "row": listing["row"], "sku": listing["sku"], "title": listing["title"],
                "reason": f"multi-unit/composite listing ({listing['unit_evidence']}); no unit-price derivation permitted",
            }
            (skipped if exact_identity_products else unmatched).append(item)
        elif len(possible_products) > 1:
            skipped.append({
                "row": listing["row"], "sku": listing["sku"], "title": listing["title"],
                "reason": "ambiguous: family has multiple OpenCart variants but spreadsheet omits a required discriminator",
            })
        else:
            unmatched.append({
                "row": listing["row"], "sku": listing["sku"], "title": listing["title"],
                "reason": "no OpenCart product has the required exact product signature",
            })
    return {"applied": applied, "skipped_or_ambiguous": skipped, "unmatched": unmatched}


def write_sql(applied: list[dict[str, object]]) -> None:
    lines = [
        "-- Generated by scripts/sync-ml-products.py from the supplied spreadsheet.",
        "-- Only high-confidence, one-sale-unit matches are included.",
        "SET NAMES utf8mb4;",
        "START TRANSACTION;",
    ]
    for match in sorted(applied, key=lambda item: int(item["product_id"])):
        source = match["source"]
        sets = [
            f"price = {float(source['price']):.4f}",
            f"quantity = {int(source['quantity'])}",
            f"status = {1 if source['status'].casefold() == 'ativo' else 0}",
            "date_modified = NOW()",
        ]
        for field in ("length", "width", "height", "weight"):
            if source[field] is not None:
                sets.append(f"{field} = {float(source[field]):.8f}")
        lines.append(f"UPDATE ws_product SET {', '.join(sets)} WHERE product_id = {int(match['product_id'])};")
        if source["description"]:
            lines.append(
                "UPDATE ws_product_description SET description = "
                f"'{sql_string(html_description(str(source['description'])))}' "
                f"WHERE product_id = {int(match['product_id'])} AND language_id = {PT_BR_LANGUAGE_ID};"
            )
    lines.append("COMMIT;")
    OUT_SQL.write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_report(audit: dict[str, object], product_count: int, listing_count: int) -> None:
    applied = audit["applied"]
    skipped = audit["skipped_or_ambiguous"]
    unmatched = audit["unmatched"]
    lines = [
        "MIIGTOOLS Mercado Livre catalog correction audit",
        f"Spreadsheet: {XLSX}",
        f"Listings audited: {listing_count}; OpenCart pt-BR products audited: {product_count}",
        f"High-confidence updates: {len(applied)}",
        f"Skipped / ambiguous listings: {len(skipped)}",
        f"Unmatched spreadsheet listings: {len(unmatched)}",
        "",
        "MATCHING RULES",
        "- A listing must sell exactly one OpenCart sale unit. Multi-unit and composite listings are never divided into a unit price.",
        "- The product family must match and all family-specific identifiers must match exactly:",
        "  bits/bedames: all inch dimensions + material grade; porta-bits: inch size;",
        "  bucha: Morse taper pair; ponta: subtype + Morse cone (+ tubular dimensions);",
        "  ER40 kits: mandrel interface; recartilhas: type + shank size.",
        "- Duplicate exact candidates use active status, then nonzero stock, then latest spreadsheet row; alternatives are reported.",
        "- PRICE is the spreadsheet PRICE exactly; tiered-pricing columns are ignored. QUANTITY is copied exactly, not multiplied.",
        "- Source STATUS maps to OpenCart status only for high-confidence exact matches. Only language_id=2 descriptions are changed.",
        "",
        "STOCK & SHIPPING FIELD MAPPING",
        "- OpenCart quantity = spreadsheet QUANTITY exactly for the matching sale unit; title pack counts never multiply inventory.",
        "- OpenCart length = SHIPPING_DEPTH (cm), width = SHIPPING_WIDTH (cm), height = SHIPPING_HEIGHT (cm), weight = SHIPPING_WEIGHT (kg).",
        "- A shipping field is overwritten only when its spreadsheet value is finite and non-negative; zero is retained as a supplied numeric value.",
        "",
        "APPLIED HIGH-CONFIDENCE MATCHES",
    ]
    for match in sorted(applied, key=lambda item: int(item["product_id"])):
        source = match["source"]
        lines.extend([
            f"product_id={match['product_id']} | {match['oc_name']} | OC SKU={match['oc_sku']}",
            f"  source row={source['row']} SKU={source['sku'] or '-'} | {source['title']}",
            f"  rationale: {match['rationale']}; confidence=high",
            f"  old={match['old']}",
            f"  new={{'price': {source['price']}, 'quantity': {source['quantity']}, 'status': {source['status']}, "
            f"'length': {source['length']}, 'width': {source['width']}, 'height': {source['height']}, 'weight': {source['weight']}}}",
            f"  changed fields={match['changes'] or 'none'}; alternative rows={match['alternative_source_rows'] or 'none'}",
        ])
    lines.extend(["", "SKIPPED / AMBIGUOUS LISTINGS"])
    lines.extend(f"row={item['row']} SKU={item['sku'] or '-'} | {item['title']} | {item['reason']}" for item in skipped)
    lines.extend(["", "UNMATCHED SPREADSHEET LISTINGS"])
    lines.extend(f"row={item['row']} SKU={item['sku'] or '-'} | {item['title']} | {item['reason']}" for item in unmatched)
    OUT_REPORT.write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    products = load_products()
    listings = load_listings()
    audit = build_audit(products, listings)
    audit["summary"] = {
        "listings_audited": len(listings),
        "products_audited": len(products),
        "high_confidence_updates": len(audit["applied"]),
        "skipped_or_ambiguous": len(audit["skipped_or_ambiguous"]),
        "unmatched": len(audit["unmatched"]),
        "multi_unit_listings_skipped": sum(listing["units"] != 1 for listing in listings),
        "matched_quantity_fields": len(audit["applied"]),
        "matched_shipping_fields_provided": sum(
            1
            for match in audit["applied"]
            for field in ("length", "width", "height", "weight")
            if match["source"][field] is not None
        ),
        "families_updated": Counter(item["source"]["signature"]["family"] for item in audit["applied"]),
    }
    OUT_JSON.write_text(json.dumps(audit, ensure_ascii=False, indent=2, default=dict), encoding="utf-8")
    write_sql(audit["applied"])
    write_report(audit, len(products), len(listings))
    print(json.dumps(audit["summary"], ensure_ascii=False, indent=2, default=dict))
    print(f"Wrote {OUT_JSON}, {OUT_SQL}, {OUT_REPORT}")


if __name__ == "__main__":
    main()
